#====== 導入所需套件 ======
import streamlit as st
import cloudscraper
import json
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.cell.cell import MergedCell
import os
import time
import math
import random  # 新增這行
import concurrent.futures
from openpyxl import Workbook

class WalletAnalyzer:
    def __init__(self):
        self.session = None
        self.last_request_time = 0
        self.min_request_interval = 2  # 最小請求間隔（秒）
        self.max_retries = 5  # 最大重試次數
        self.base_delay = 3  # 基礎延遲時間（秒）
        self.retry_count = 0  # 重試計數器
        self.initialize_session()

    def initialize_session(self):
        """初始化會話並進行必要的準備工作"""
        while self.retry_count < 3:  # 最多嘗試初始化3次
            try:
                self.session = cloudscraper.create_scraper(
                    browser={
                        'browser': 'chrome',
                        'platform': 'windows',
                        'desktop': True
                    },
                    debug=True,  # 開啟調試模式
                    defaults={
                        'timeout': 30,
                        'allow_redirects': True
                    }
                )
                
                # 更新請求頭，模擬更真實的瀏覽器行為
                self.headers = {
                    'authority': 'gmgn.ai',
                    'accept': '*/*',
                    'accept-language': 'en-US,en;q=0.9,zh-TW;q=0.8,zh;q=0.7',
                    'cache-control': 'no-cache',
                    'content-type': 'application/json',
                    'dnt': '1',
                    'origin': 'https://gmgn.ai',
                    'pragma': 'no-cache',
                    'referer': 'https://gmgn.ai/',
                    'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
                    'sec-ch-ua-mobile': '?0',
                    'sec-ch-ua-platform': '"Windows"',
                    'sec-fetch-dest': 'empty',
                    'sec-fetch-mode': 'cors',
                    'sec-fetch-site': 'same-origin',
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
                    'x-requested-with': 'XMLHttpRequest'
                }
    
                print("嘗試訪問首頁...")
                # 首先訪問首頁
                response = self.session.get(
                    'https://gmgn.ai/',
                    headers=self.headers,
                    timeout=30,
                    allow_redirects=True
                )
                
                print(f"首頁響應狀態碼: {response.status_code}")
                print(f"首頁響應頭: {dict(response.headers)}")
                
                if response.status_code == 200:
                    print("Session initialized successfully")
                    self.retry_count = 0  # 重置重試計數
                    time.sleep(2)
                    return
                else:
                    print(f"初始化失敗，狀態碼: {response.status_code}")
                    print(f"響應內容: {response.text[:500]}")  # 只打印前500個字符
                    
                    if response.status_code == 403:
                        print("檢測到 Cloudflare 防護，等待更長時間...")
                        time.sleep(10)  # 遇到 403 時等待更長時間
                    
                    raise Exception(f"Initialization failed with status code: {response.status_code}")
                    
            except Exception as e:
                self.retry_count += 1
                print(f"Session initialization attempt {self.retry_count} failed: {str(e)}")
                if self.retry_count < 3:
                    sleep_time = 5 * self.retry_count  # 逐漸增加等待時間
                    print(f"Waiting {sleep_time} seconds before retry...")
                    time.sleep(sleep_time)
                else:
                    print("Failed to initialize session after 3 attempts")
                    # 不要拋出異常，而是使用備用方案
                    print("Using backup initialization...")
                    self.session = cloudscraper.create_scraper(
                        defaults={
                            'timeout': 30,
                            'allow_redirects': True
                        }
                    )
                    self.retry_count = 0
                    return

    def enforce_rate_limit(self):
        """強制執行請求速率限制"""
        current_time = time.time()
        time_since_last_request = current_time - self.last_request_time
        if time_since_last_request < self.min_request_interval:
            sleep_time = self.min_request_interval - time_since_last_request + random.uniform(0.1, 0.5)
            time.sleep(sleep_time)
        self.last_request_time = time.time()

    def make_request(self, url, method='get', params=None, data=None, retry_count=0):
        """發送請求並處理可能的錯誤"""
        self.enforce_rate_limit()

        try:
            if method.lower() == 'get':
                response = self.session.get(
                    url,
                    params=params,
                    headers=self.headers,
                    timeout=30
                )
            else:
                response = self.session.post(
                    url,
                    params=params,
                    json=data,
                    headers=self.headers,
                    timeout=30
                )

            # 打印響應信息用於調試
            print(f"Response status: {response.status_code}")
            print(f"Response headers: {dict(response.headers)}")

            if response.status_code == 403:
                print("Received 403 error, reinitializing session...")
                self.initialize_session()
                if retry_count < self.max_retries:
                    time.sleep(5)  # 等待5秒後重試
                    return self.make_request(url, method, params, data, retry_count + 1)
                else:
                    raise Exception("Maximum retries reached for 403 error")

            response.raise_for_status()
            return response.json()

        except Exception as e:
            if retry_count >= self.max_retries:
                raise Exception(f"Maximum retries reached: {str(e)}")

            # 計算延遲時間（指數退避）
            delay = min(300, self.base_delay * (2 ** retry_count) + random.uniform(0, 1))
            print(f"Request failed, retrying in {delay:.2f} seconds... (Attempt {retry_count + 1}/{self.max_retries})")
            time.sleep(delay)

            return self.make_request(url, method, params, data, retry_count + 1)
            
    #====== 獲取代幣首次購買市值方法 ======
    def get_token_first_buy_marketcap(self, wallet_address, token_address):
        """獲取代幣首次購買市值方法"""
        max_retries = 3
        retry_delay = 10  # 10秒重試

        for attempt in range(max_retries):
            try:
                # 使用線程池並行請求交易歷史和代幣信息
                with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
                    # 提交兩個任務
                    activities_future = executor.submit(
                        self.scraper.get,
                        "https://gmgn.ai/defi/quotation/v1/wallet_token_activity/sol",
                        params={"wallet": wallet_address, "token": token_address, "limit": 50},
                        headers=self.headers
                    )
                    
                    token_info_future = executor.submit(
                        self.scraper.get,
                        f"https://gmgn.ai/api/v1/token_info/sol/{token_address}",
                        headers=self.headers
                    )

                    # 獲取結果
                    response = activities_future.result()
                    token_info_response = token_info_future.result()

                data = response.json()
                token_info = token_info_response.json()
                
                if not isinstance(data, dict) or 'data' not in data or 'activities' not in data['data']:
                    if attempt < max_retries - 1:
                        print(f"重試獲取交易歷史 ({token_address}) - 嘗試 {attempt + 1}/{max_retries}")
                        time.sleep(retry_delay)
                        continue
                    return None, "無交易記錄"

                first_buy = None
                for activity in reversed(data['data']['activities']):
                    if activity['event_type'] == 'buy':
                        first_buy = activity
                        break
                
                if not first_buy:
                    return None, "無買入記錄"

                if not isinstance(token_info, dict) or 'data' not in token_info:
                    if attempt < max_retries - 1:
                        print(f"重試獲取代幣信息 ({token_address}) - 嘗試 {attempt + 1}/{max_retries}")
                        time.sleep(retry_delay)
                        continue
                    return None, "無法獲取代幣信息"

                try:
                    # 確保價格和供應量是純數值
                    price_usd = float(f"{first_buy['price_usd']:.16f}")
                    total_supply = float(token_info['data']['total_supply'])
                    # 計算市值並四捨五入為整數
                    market_cap = round(total_supply * price_usd)

                    print(f"\n代幣 {token_address} 的計算:")
                    print(f"首購價格: {price_usd}")
                    print(f"總供應量: {total_supply:,.0f}")
                    print(f"計算出的市值: ${market_cap:,}")

                    # 直接返回純數值市值和時間戳
                    return market_cap, first_buy['timestamp']
                
                except (ValueError, TypeError) as e:
                    print(f"市值計算過程中發生錯誤: {str(e)}")
                    if attempt < max_retries - 1:
                        continue
                    return None, f"市值計算錯誤: {str(e)}"

            except Exception as e:
                if attempt < max_retries - 1:
                    print(f"重試處理代幣 ({token_address}) - 嘗試 {attempt + 1}/{max_retries}")
                    time.sleep(retry_delay)
                    continue
                print(f"計算市值時發生錯誤 ({token_address}): {str(e)}")
                return None, f"錯誤: {str(e)}"

        return None, "重試次數已達上限"
    #====== 分析交易模式，計算快速交易比例 ======
    def analyze_trading_pattern(self, wallet_address, token_address):
        """
        分析交易模式，計算快速交易比例
        """
        max_retries = 3
        retry_delay = 10

        for attempt in range(max_retries):
            try:
                # 首先獲取代幣資訊
                token_info_response = self.scraper.get(
                    f"https://gmgn.ai/api/v1/token_info/sol/{token_address}",
                    headers=self.headers
                )
                token_info = token_info_response.json()
                
                # 獲取代幣符號
                token_symbol = None
                if token_info and 'data' in token_info:
                    token_symbol = token_info['data'].get('symbol')
                
                # 獲取交易活動
                response = self.scraper.get(
                    "https://gmgn.ai/defi/quotation/v1/wallet_token_activity/sol",
                    params={"wallet": wallet_address, "token": token_address, "limit": 50},
                    headers=self.headers
                )
                
                data = response.json()
                
                if not isinstance(data, dict) or 'data' not in data or 'activities' not in data['data']:
                    if attempt < max_retries - 1:
                        continue
                    return None

                activities = data['data']['activities']
                
                # 交易統計
                stats = {
                    'total_trades': 0,          
                    'quick_trades': 0,          
                    'holding_times': [],
                    'quick_trade_details': []    
                }
                
                # 按時間排序
                activities.sort(key=lambda x: int(x['timestamp']))
                
                # 分析買賣配對
                buy_time = None
                buy_market_cap = None
                buy_price = None
                
                for activity in activities:
                    current_time = int(activity['timestamp'])
                    
                    if activity['event_type'] == 'buy':
                        buy_time = current_time
                        # 計算買入時的市值
                        if 'price_usd' in activity and token_info and 'data' in token_info:
                            price_usd = float(f"{activity['price_usd']:.16f}")
                            total_supply = float(token_info['data']['total_supply'])
                            buy_market_cap = round(total_supply * price_usd)
                        buy_price = float(activity['price_usd']) if 'price_usd' in activity else None
                        
                    elif activity['event_type'] == 'sell' and buy_time and buy_price is not None:
                        holding_time = current_time - buy_time
                        stats['holding_times'].append(holding_time)
                        stats['total_trades'] += 1
                        
                        # 計算收益率
                        sell_price = float(activity['price_usd']) if 'price_usd' in activity else None
                        profit_rate = ((sell_price - buy_price) / buy_price) if sell_price and buy_price else None
                        
                        # 統計1分鐘內的快速交易
                        if holding_time < 60:  # 小於1分鐘
                            stats['quick_trades'] += 1
                            # 添加快速交易詳情
                            stats['quick_trade_details'].append({
                                'token_symbol': token_symbol or '未知代幣',
                                'buy_market_cap': buy_market_cap,
                                'profit_rate': profit_rate,
                                'interval': holding_time
                            })
                        
                        buy_time = None
                        buy_market_cap = None
                        buy_price = None

                if stats['total_trades'] > 0:
                    quick_trade_ratio = stats['quick_trades'] / stats['total_trades']
                    
                    return {
                        'quick_trade_ratio': quick_trade_ratio,
                        'total_trades': stats['total_trades'],
                        'quick_trades': stats['quick_trades'],
                        'quick_trade_details': stats['quick_trade_details']
                    }

                return None

            except Exception as e:
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    continue
                print(f"分析交易模式時發生錯誤 ({token_address}): {str(e)}")
                return None

        return None

    #====== 獲取交易記錄方法 ======
    def fetch_transactions(self, wallet_address):
        """獲取錢包交易記錄"""
        st.write("開始獲取交易記錄...")
        transactions = []
        cursor = None
        current_time = datetime.now()
        thirty_days_ago = current_time - timedelta(days=30)
        
        # 添加進度條
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        retry_count = 0
        max_retries = 3
        
        while retry_count < max_retries:
            try:
                status_text.text(f"正在嘗試第 {retry_count + 1} 次獲取數據...")
                
                url = f"https://gmgn.ai/api/v1/wallet_holdings/sol/{wallet_address}"
                params = {
                    "limit": "50",
                    "orderby": "last_active_timestamp",
                    "direction": "desc",
                    "showsmall": "true",
                    "sellout": "true",
                    "tx30d": "true"
                }
                
                if cursor:
                    params["cursor"] = cursor
                
                # 設置較短的超時時間
                response = self.session.get(
                    url,
                    params=params,
                    headers=self.headers,
                    timeout=10
                )
                
                if response.status_code != 200:
                    raise Exception(f"請求失敗，狀態碼: {response.status_code}")
                
                data = response.json()
                
                if not data.get('data') or not data['data'].get('holdings'):
                    status_text.text("未找到交易數據")
                    break
                
                for holding in data['data']['holdings']:
                    # 處理數據...（保持原有的處理邏輯）
                    last_active_timestamp = holding.get('last_active_timestamp')
                    if not isinstance(last_active_timestamp, (int, float)):
                        continue
                    
                    last_active = datetime.fromtimestamp(last_active_timestamp)
                    if last_active < thirty_days_ago:
                        status_text.text("已獲取所有30天內的交易記錄")
                        return transactions
                    
                    total_profit_pnl = holding.get('total_profit_pnl')
                    if total_profit_pnl is not None:
                        try:
                            total_profit_pnl = float(total_profit_pnl)
                            transactions.append({
                                '時間戳': last_active.strftime('%Y-%m-%d %H:%M:%S'),
                                '代幣名稱': holding.get('token', {}).get('symbol', 'Unknown'),
                                '收益率': total_profit_pnl,
                                '交易類型': '獲利' if total_profit_pnl > 0 else '虧損',
                                'token': holding.get('token', {})
                            })
                        except (ValueError, TypeError) as e:
                            st.warning(f"處理交易數據時發生錯誤: {str(e)}")
                            continue
                
                if 'next' in data['data'] and data['data']['next']:
                    cursor = data['data']['next']
                    progress_bar.progress(min(len(transactions) / 100, 1.0))  # 更新進度條
                    time.sleep(1)  # 添加延遲避免請求過快
                else:
                    break
                
                # 成功獲取數據，重置重試計數
                retry_count = 0
                
            except Exception as e:
                retry_count += 1
                if retry_count >= max_retries:
                    st.error(f"獲取交易記錄失敗: {str(e)}")
                    return []
                
                wait_time = 5 * retry_count
                status_text.text(f"請求失敗，等待 {wait_time} 秒後重試...")
                time.sleep(wait_time)
        
        status_text.text(f"成功獲取 {len(transactions)} 筆交易記錄")
        progress_bar.empty()  # 清除進度條
        return transactions

    #====== 分析交易數據方法 ======
    def analyze_transactions(self, transactions):
        if not transactions:
            return None

        profit_trades = [t for t in transactions if t['收益率'] > 0]
        loss_trades = [t for t in transactions if t['收益率'] < 0]

        avg_profit = sum(t['收益率'] for t in profit_trades) / len(profit_trades) if profit_trades else 0
        avg_loss = sum(t['收益率'] for t in loss_trades) / len(loss_trades) if loss_trades else -1

        risk_reward_ratio = round(abs(avg_profit / abs(avg_loss)), 2) if avg_loss != 0 else 0

        return {
            '總交易次數': len(transactions),
            '獲利次數': len(profit_trades),
            '虧損次數': len(loss_trades),
            '平均獲利': round(avg_profit * 100, 2),
            '平均虧損': round(avg_loss * 100, 2),
            '最大獲利': round(max((t['收益率'] for t in profit_trades), default=0) * 100, 2),
            '最大虧損': round(min((t['收益率'] for t in loss_trades), default=0) * 100, 2),
            '勝率': round(len(profit_trades) / len(transactions) * 100, 2) if transactions else 0,
            '盈虧比': risk_reward_ratio
        }

    #====== 分析代幣收益情況方法 ======
    def analyze_tokens_by_profit(self, wallet_address, transactions):
        """分析代幣收益情況方法"""
        profit_tokens = {}
        loss_tokens = {}
        failed_tokens = []
        total_quick_trades = 0
        bot_trades = []  # 保存快速交易詳情
        rug_count = 0  # 用於計算Rug盤數量
        total_tokens = 0  # 用於計算代幣總數
        processed_tokens = set()  # 用於追蹤已處理的代幣
        
        # 使用實際的交易記錄數量
        total_trades = len(transactions)
        
        # 獲取5倍爆擊率
        try:
            response = self.scraper.get(
                f"https://gmgn.ai/defi/quotation/v1/smartmoney/sol/walletNew/{wallet_address}",
                params={"period": "30d"},
                headers=self.headers
            )
            data = response.json()
            print(f"API 返回5倍爆擊數據: {data}")  # 保留調試輸出
            
            if 'data' in data and isinstance(data['data'], dict):
                pnl_gt_5x_num = float(data['data'].get('pnl_gt_5x_num', 0))
                
                # 使用 total_trades 作為分母
                # total_trades 是我們從交易記錄中計算出來的總交易次數
                total_trades = total_trades  # 使用傳入的 total_trades 參數
                
                print(f"5倍爆擊次數: {pnl_gt_5x_num}")
                print(f"總交易次數（自行計算）: {total_trades}")
                
                if total_trades > 0 and isinstance(pnl_gt_5x_num, (int, float)):
                    five_x_rate = pnl_gt_5x_num / total_trades
                else:
                    print(f"無效的數據: pnl_gt_5x_num = {pnl_gt_5x_num}, total_trades = {total_trades}")
                    five_x_rate = 0.0
            else:
                print(f"API返回數據格式異常: {data}")
                five_x_rate = 0.0
        except Exception as e:
            print(f"獲取5倍爆擊率時發生錯誤: {str(e)}")
            five_x_rate = 0.0

        print(f"最終計算的5倍爆擊率: {five_x_rate*100:.2f}%")  # 保留最終結果輸出
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            future_to_token = {}
            
            # 首先計算 Rug 盤比例
            for tx in transactions:
                if 'token' not in tx or 'address' not in tx['token']:
                    continue
                    
                token_address = tx['token']['address']
                
                # 如果已經處理過這個代幣，跳過
                if token_address in processed_tokens:
                    continue
                
                processed_tokens.add(token_address)
                total_tokens += 1
                
                # 檢查是否為 Rug 盤
                if tx['token'].get('is_show_alert') is True:
                    rug_count += 1
                
                try:
                    # 同時獲取市值和交易模式
                    market_cap_future = executor.submit(
                        self.get_token_first_buy_marketcap,
                        wallet_address,
                        token_address
                    )
                    
                    trading_pattern_future = executor.submit(
                        self.analyze_trading_pattern,
                        wallet_address,
                        token_address
                    )
                    
                    future_to_token[market_cap_future] = (token_address, tx, trading_pattern_future)
                except Exception as e:
                    print(f"提交任務時發生錯誤 ({token_address}): {str(e)}")
                    continue
            
            # 處理所有完成的任務
            for market_cap_future in concurrent.futures.as_completed(future_to_token):
                token_address, tx, trading_pattern_future = future_to_token[market_cap_future]
                
                try:
                    market_cap, timestamp = market_cap_future.result()
                    trading_pattern = trading_pattern_future.result()
                    
                    # 處理交易模式數據
                    if trading_pattern and isinstance(trading_pattern, dict):
                        quick_trades = trading_pattern.get('quick_trades', 0)
                        total_quick_trades += quick_trades
                        
                        # 處理快速交易詳情
                        if trading_pattern.get('quick_trade_details'):
                            for detail in trading_pattern['quick_trade_details']:
                                if isinstance(detail, dict):
                                    detail['buy_market_cap'] = market_cap
                                    bot_trades.append(detail)
                    
                    # 只在市值有效時添加到代幣列表
                    if market_cap is not None and not isinstance(market_cap, str) and isinstance(market_cap, (int, float)):
                        token_data = {
                            'symbol': tx.get('代幣名稱', 'Unknown'),
                            'market_cap': market_cap,
                            'timestamp': timestamp,
                            'profit_rate': tx.get('收益率', 0)
                        }
                        
                        if tx.get('收益率', 0) > 0:
                            profit_tokens[token_address] = token_data
                        else:
                            loss_tokens[token_address] = token_data
                    else:
                        failed_tokens.append((token_address, tx))
                    
                except Exception as e:
                    print(f"處理代幣 {token_address} 時發生錯誤: {str(e)}")
                    failed_tokens.append((token_address, tx))
                    continue

            # 重試失敗的代幣
            if failed_tokens:
                print(f"\n開始重試 {len(failed_tokens)} 個失敗的代幣...")
                retry_tokens = failed_tokens.copy()
                failed_tokens.clear()
                
                time.sleep(10)  # 等待10秒後重試
                
                for token_address, tx in retry_tokens:
                    try:
                        market_cap, timestamp = self.get_token_first_buy_marketcap(wallet_address, token_address)
                        trading_pattern = self.analyze_trading_pattern(wallet_address, token_address)
                        
                        if market_cap is not None and not isinstance(market_cap, str) and isinstance(market_cap, (int, float)):
                            token_data = {
                                'symbol': tx.get('代幣名稱', 'Unknown'),
                                'market_cap': market_cap,
                                'timestamp': timestamp,
                                'profit_rate': tx.get('收益率', 0)
                            }
                            
                            if tx.get('收益率', 0) > 0:
                                profit_tokens[token_address] = token_data
                            else:
                                loss_tokens[token_address] = token_data
                                
                            if trading_pattern and isinstance(trading_pattern, dict):
                                total_quick_trades += trading_pattern.get('quick_trades', 0)
                                if trading_pattern.get('quick_trade_details'):
                                    for detail in trading_pattern['quick_trade_details']:
                                        if isinstance(detail, dict):
                                            detail['buy_market_cap'] = market_cap
                                            bot_trades.append(detail)
                        else:
                            failed_tokens.append((token_address, tx))
                    except Exception as e:
                        print(f"重試處理代幣 {token_address} 時發生錯誤: {str(e)}")
                        failed_tokens.append((token_address, tx))

        # 計算最終的比率
        bot_ratio = total_quick_trades / total_trades * 100 if total_trades > 0 else 0
        rug_ratio = rug_count / total_tokens if total_tokens > 0 else 0

        # 打印驗證信息
        print(f"\nBot交易統計驗證：")
        print(f"總交易次數（實際）: {total_trades}")
        print(f"快速交易次數: {total_quick_trades}")
        print(f"計算出的Bot機率: {bot_ratio:.4f}%")
        print(f"Rug盤比例: {rug_ratio*100:.4f}%")
        print(f"5倍爆擊率: {five_x_rate*100:.4f}%")

        return {
            'profit_tokens': profit_tokens,
            'loss_tokens': loss_tokens,
            'rug_ratio': rug_ratio,
            'five_x_rate': five_x_rate,
            'quick_trade_ratio': bot_ratio,
            'quick_trade_details': bot_trades,
            'total_trades': total_trades,
            'total_quick_trades': total_quick_trades
        }
    #====== 保存到Excel方法 ======
    def save_to_excel(self, wallet_address, transactions, analysis, token_analysis, advanced_analysis):
        if not transactions:
            print(f"錢包 {wallet_address} 沒有交易記錄")
            return

        output_directory = "/Users/liuian/Desktop/聰明錢分析/錢包操作分析"
        os.makedirs(output_directory, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "交易分析"

        # 合併儲存格並添加標題
        ws.merge_cells('A1:D1')
        ws['A1'] = '錢包地址'
        ws['A1'].font = Font(name='Arial', size=11, bold=True)  # 設置粗體
        
        # 合併儲存格並添加錢包地址
        ws.merge_cells('A2:D2')
        ws['A2'] = wallet_address
        ws['A2'].font = Font(name='Arial', size=11, bold=True)  # 設置粗體
        
        ws.append([])  # 空行
        
        # 合併儲存格並添加分析結果標題
        ws.merge_cells('A3:D3')
        ws['A3'] = '基礎分析結果'
        ws['A3'].font = Font(name='Arial', size=11, bold=True)  # 設置粗體
        
        # 交易分析結果
        current_row = 4  # 從第4行開始添加分析結果
        
        # 將百分比值除以100以得到正確的顯示
        percentage_fields = ['平均獲利', '平均虧損', '最大獲利', '最大虧損', '勝率']
        for key, value in analysis.items():
            if key in percentage_fields:
                # 將百分比值除以100
                adjusted_value = value / 100
                ws.cell(row=current_row, column=1, value=key)
                ws.cell(row=current_row, column=2, value=adjusted_value)
                current_row += 1
            else:
                ws.cell(row=current_row, column=1, value=key)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1

        # 添加分析指標到分析結果中
        ws.cell(row=current_row, column=1, value='Bot機率')
        ws.cell(row=current_row, column=2, value=token_analysis['quick_trade_ratio']/100)  # 轉換為小數以顯示為百分比
        current_row += 1
        
        # 新增 Rug 盤比例
        ws.cell(row=current_row, column=1, value='Rug盤比例')
        ws.cell(row=current_row, column=2, value=token_analysis['rug_ratio'])  # 已經是小數形式
        current_row += 1
        
        # 新增 5倍爆擊率
        ws.cell(row=current_row, column=1, value='5倍爆擊率')
        ws.cell(row=current_row, column=2, value=token_analysis['five_x_rate'])  # 已經是小數形式
        current_row += 1

        current_row += 1  # 空行

        # 添加固定市值區間分析
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value='固定市值區間分析')
        ws.cell(row=current_row, column=1).font = Font(name='Arial', size=11, bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        # 添加框線
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1

        # 添加固定區間分析表頭
        headers = ['區間', '交易次數', '勝率', '綜合得分']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(name='Arial', size=11, bold=True)
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            ws.cell(row=current_row, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        fixed_range_filter_row = current_row
        current_row += 1

        # 添加固定區間數據
        for range_name, metrics in advanced_analysis['fixed_ranges'].items():
            ws.cell(row=current_row, column=1, value=range_name)
            ws.cell(row=current_row, column=2, value=metrics['交易次數'])
            ws.cell(row=current_row, column=3, value=metrics['勝率'] / 100)  # 轉換為小數以顯示為百分比
            ws.cell(row=current_row, column=4, value=metrics['綜合得分'])
            # 添加框線
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            current_row += 1

        current_row += 1  # 空行

        # 添加動態最佳區間分析
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value='動態市值區間分析')
        ws.cell(row=current_row, column=1).font = Font(name='Arial', size=11, bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        # 添加框線
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        current_row += 1

        if advanced_analysis['dynamic_range']['range']:
            best_range = advanced_analysis['dynamic_range']['range']
            metrics = advanced_analysis['dynamic_range']['metrics']
            
            # 格式化區間顯示
            range_str = f"${best_range[0]:,.0f} - ${best_range[1]:,.0f}"
            
            ws.merge_cells(f'A{current_row}:D{current_row}')
            ws.cell(row=current_row, column=1, value=f'最佳交易市值區間: {range_str}')
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            # 添加框線
            for col in range(1, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            current_row += 1

            # 添加詳細指標
            metrics_display = [
                ('交易次數', metrics['交易次數']),
                ('勝率', metrics['勝率'] / 100),
                ('平均獲利', metrics['平均獲利'] / 100),
                ('平均虧損', metrics['平均虧損'] / 100),
                ('盈虧比', metrics['盈虧比']),
                ('區間內代幣數', metrics['區間內代幣數']),
                ('綜合得分', metrics['綜合得分'])
            ]

            for label, value in metrics_display:
                ws.cell(row=current_row, column=1, value=label)
                cell = ws.cell(row=current_row, column=2, value=value)
                if label in ['勝率', '平均獲利', '平均虧損']:
                    cell.number_format = '0.00%'
                elif label in ['盈虧比', '綜合得分']:
                    cell.number_format = '0.00'
                # 添加框線
                for col in range(1, 5):
                    ws.cell(row=current_row, column=col).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                current_row += 1

        current_row += 1  # 空行

        # Bot 交易記錄標題
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value='快速交易紀錄')
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # 為合併儲存格的所有單元格設置填充顏色
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        current_row += 1
        
        # Bot 交易記錄表頭
        headers = ['代幣名稱', '首購市值', '收益率', '交易間隔(秒)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(name='Arial', size=11, bold=True)
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        bot_trades_filter_row = current_row
        current_row += 1
        
        # 添加快速交易紀錄數據
        quick_trade_details = token_analysis.get('quick_trade_details', [])
        for trade in quick_trade_details:
            ws.cell(row=current_row, column=1, value=trade['token_symbol'])
            ws.cell(row=current_row, column=2, value=trade['buy_market_cap'])
            
            # 設置收益率
            if trade['profit_rate'] is not None:
                ws.cell(row=current_row, column=3, value=trade['profit_rate'])
            else:
                ws.cell(row=current_row, column=3, value='N/A')
                
            ws.cell(row=current_row, column=4, value=trade['interval'])
            current_row += 1
        
        current_row += 1  # 空行

        # 獲利代幣詳情
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value='獲利代幣詳情')
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # 為合併儲存格的所有單元格設置填充顏色
        for col in range(1, 5):  # A到D列
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        current_row += 1
        
        headers = ['代幣符號', '首購市值', '收益率', '首購時間']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(name='Arial', size=11, bold=True)  # 加粗表頭
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # 添加填充色
        profit_filter_row = current_row
        current_row += 1
        
        # 添加獲利代幣數據
        for token in token_analysis['profit_tokens'].values():
            try:
                if token['timestamp'] and token['timestamp'] != "無買入記錄" and token['timestamp'] != "無交易記錄":
                    timestamp = datetime.fromtimestamp(int(token['timestamp'])).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    timestamp = 'N/A'
            except (ValueError, TypeError):
                timestamp = 'N/A'

            ws.cell(row=current_row, column=1, value=token['symbol'])
            ws.cell(row=current_row, column=2, value=round(token['market_cap']) if token['market_cap'] else 'N/A')
            ws.cell(row=current_row, column=3, value=token['profit_rate'])
            ws.cell(row=current_row, column=4, value=timestamp)
            current_row += 1
        
        current_row += 1  # 空行
        
        # 虧損代幣詳情
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value='虧損代幣詳情')
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # 為合併儲存格的所有單元格設置填充顏色
        for col in range(1, 5):  # A到D列
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        current_row += 1
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(name='Arial', size=11, bold=True)  # 加粗表頭
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # 添加填充色
        loss_filter_row = current_row
        current_row += 1
        
        # 添加虧損代幣數據
        for token in token_analysis['loss_tokens'].values():
            try:
                if token['timestamp'] and token['timestamp'] != "無買入記錄" and token['timestamp'] != "無交易記錄":
                    timestamp = datetime.fromtimestamp(int(token['timestamp'])).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    timestamp = 'N/A'
            except (ValueError, TypeError):
                timestamp = 'N/A'

            ws.cell(row=current_row, column=1, value=token['symbol'])
            ws.cell(row=current_row, column=2, value=round(token['market_cap']) if token['market_cap'] else 'N/A')
            ws.cell(row=current_row, column=3, value=token['profit_rate'])
            ws.cell(row=current_row, column=4, value=timestamp)
            current_row += 1
        
        current_row += 1  # 空行

        # 交易記錄標題
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws.cell(row=current_row, column=1, value='交易記錄')
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # 為合併儲存格的所有單元格設置填充顏色
        for col in range(1, 5):  # A到D列
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        current_row += 1
        
        trade_headers = ['時間戳', '代幣名稱', '收益率', '交易類型']
        for col, header in enumerate(trade_headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(name='Arial', size=11, bold=True)  # 加粗表頭
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # 添加填充色
        trade_filter_row = current_row
        current_row += 1
        
        # 添加交易記錄數據
        for tx in transactions:
            ws.cell(row=current_row, column=1, value=tx['時間戳'])
            ws.cell(row=current_row, column=2, value=tx['代幣名稱'])
            ws.cell(row=current_row, column=3, value=tx['收益率'])
            ws.cell(row=current_row, column=4, value=tx['交易類型'])
            current_row += 1


        # 更新格式化參數
        self.format_excel(
            ws,
            profit_filter_row,
            loss_filter_row,
            trade_filter_row,
            current_row,
            bot_trades_filter_row
        )

        # 設置工作表的視圖縮放比例為 120%
        ws.sheet_view.zoomScale = 120

        # 取得最佳區間範圍
        best_range = advanced_analysis['dynamic_range']['range']
        if best_range:
            lower_cap, upper_cap = best_range
            # 格式化最佳區間字串
            best_range_str = f"${lower_cap:,.0f} - ${upper_cap:,.0f}"
        else:
            best_range_str = "無最佳區間"

        # 生成新的檔案名稱格式
        address_prefix = wallet_address[:3]  # 取前3個字
        filename = os.path.join(
            output_directory, 
            f"{address_prefix}哥（最佳區間：{best_range_str}）.xlsx"
        )

        wb.save(filename)
        print(f"分析結果已保存到 {filename}")


    #====== Excel格式化方法 ======
    def format_excel(self, ws, profit_filter_row, loss_filter_row, trade_filter_row, max_row, bot_trades_filter_row):
        """格式化 Excel 工作表"""
        # 定義樣式
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        filter_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=11)
        bold_font = Font(name='Arial', size=11, bold=True)
        title_font = Font(name='Arial', size=14, bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 設置列寬
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 22

        # 設置主標題
        title_texts = ['錢包地址', '基礎分析結果', '固定市值區間分析', '動態市值區間分析', '獲利代幣詳情', '虧損代幣詳情', '交易記錄', '快速交易紀錄']
        
        # 設置所有可能的表頭組合
        header_combinations = [
            ['區間', '交易次數', '勝率', '綜合得分'],  # 固定區間分析表頭
            ['代幣符號', '首購市值', '收益率', '首購時間'],  # 代幣詳情表頭
            ['時間戳', '代幣名稱', '收益率', '交易類型'],  # 交易記錄表頭
            ['代幣名稱', '首購市值', '收益率', '交易間隔(秒)']  # 快速交易紀錄表頭（已更新）
        ]

        # 找到並設置所有標題
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=1)
            
            # 處理主標題
            if cell.value in title_texts:
                ws.row_dimensions[row].height = 30
                cell.font = title_font
                for col in range(1, 5):
                    current_cell = ws.cell(row=row, column=col)
                    current_cell.font = title_font
                    current_cell.fill = header_fill
                    current_cell.border = border
                    current_cell.alignment = center_alignment
            
            # 處理表頭行
            headers = []
            for col in range(1, 5):
                headers.append(ws.cell(row=row, column=col).value)
            
            # 檢查是否為任何一個表頭組合
            if headers in header_combinations:
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.font = filter_font  # 使用粗體字型
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_alignment
                ws.row_dimensions[row].height = 20

        # 設置前三列的格式
        for col in ['A1', 'B1', 'C1', 'D1', 'A2', 'B2', 'C2', 'D2', 'A3', 'B3', 'C3', 'D3']:
            ws[col].border = border
            ws[col].alignment = center_alignment

        # 為所有單元格設置基本格式
        for row in ws.iter_rows(min_row=1, max_row=max_row):
            for cell in row:
                # 基本格式
                cell.border = border
                cell.alignment = center_alignment
                
                # 如果不是標題行或表頭行，設置普通字體
                row_headers = [ws.cell(row=cell.row, column=i).value for i in range(1, 5)]
                if (cell.row not in [1, 3] and
                    cell.value not in title_texts and
                    row_headers not in header_combinations):
                    cell.font = normal_font

                # 檢查是否為要設置百分比格式的行
                cell_value_in_col_a = ws.cell(row=cell.row, column=1).value
                percentage_keywords = [
                    '平均獲利', '平均虧損', '最大獲利', '最大虧損', 
                    '勝率', 'Bot機率', 'Rug盤比例', '5倍爆擊率'
                ]
                
                try:
                    # 處理百分比格式
                    if cell_value_in_col_a in percentage_keywords and cell.column == 2:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'
                    # 收益率列特殊處理
                    elif cell.column == 3 and isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00%'
                    # 首購市值列特殊處理
                    elif cell.column == 2 and isinstance(cell.value, str) and cell.value.startswith('$'):
                        number_str = cell.value.replace('$', '').replace(',', '')
                        try:
                            number_value = round(float(number_str))
                            cell.value = number_value
                            cell.number_format = '#,##0'
                        except ValueError:
                            continue
                    # 其他數字格式
                    elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        if '盈虧比' in str(ws.cell(row=cell.row, column=1).value):
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'
                except AttributeError:
                    continue

        # Bot 交易記錄的特殊格式處理
        for row in range(bot_trades_filter_row, max_row + 1):
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                
                # 如果是時間列（第2和第3列），設置日期時間格式
                if col in [2, 3] and row > bot_trades_filter_row:
                    if isinstance(cell.value, str):
                        try:
                            # 如果是字符串格式的日期時間，轉換為datetime對象
                            cell.value = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
                            cell.number_format = 'mm-dd hh:mm:ss'
                        except ValueError:
                            continue

        # 取消總交易次數的填充色
        for row in ws.iter_rows(min_row=1, max_row=max_row):
            cell = row[0]
            if cell.value == "總交易次數":
                cell.fill = PatternFill(fill_type=None)
                if len(row) > 1:
                    row[1].fill = PatternFill(fill_type=None)

        # 調整行高
        for row in range(1, max_row + 1):
            if ws.cell(row=row, column=1).value not in title_texts:
                ws.row_dimensions[row].height = 20


    #====== 動態尋找最佳市值交易區間 ======
    def find_optimal_market_cap_range(self, wallet_address, transactions, token_analysis):
        """動態尋找最佳市值交易區間"""
        # 準備數據
        token_trades = []
        
        # 收集所有有效的交易數據
        for token_address, token_data in {**token_analysis['profit_tokens'], **token_analysis['loss_tokens']}.items():
            market_cap = token_data.get('market_cap')
            if market_cap:  # 只收集有市值數據的交易
                for tx in transactions:
                    if 'token' in tx and 'address' in tx['token'] and tx['token']['address'] == token_address:
                        token_trades.append({
                            'market_cap': market_cap,
                            'profit_rate': tx['收益率']
                        })
        
        if not token_trades:
            return None
        
        # 按市值排序所有交易
        token_trades.sort(key=lambda x: x['market_cap'])
        
        # 獲取市值範圍
        min_market_cap = token_trades[0]['market_cap']
        max_market_cap = token_trades[-1]['market_cap']
        
        # 設置搜索參數
        MIN_TRADES = 5  # 最小交易數量要求
        win_rate_weight = 0.35      # 勝率權重
        profit_ratio_weight = 0.35  # 收益率權重
        trade_count_weight = 0.30   # 交易數量權重
        
        best_score = 0
        best_range = None
        best_metrics = None
        
        # 使用滑動窗口尋找最佳區間
        for lower_idx in range(len(token_trades)):
            lower_cap = token_trades[lower_idx]['market_cap']
            
            for upper_idx in range(lower_idx + MIN_TRADES - 1, len(token_trades)):
                upper_cap = token_trades[upper_idx]['market_cap']
                
                # 獲取區間內的所有交易
                range_trades = token_trades[lower_idx:upper_idx + 1]
                
                if len(range_trades) < MIN_TRADES:
                    continue
                    
                # 計算區間指標
                total_trades = len(range_trades)
                profit_trades = [t for t in range_trades if t['profit_rate'] > 0]
                win_rate = (len(profit_trades) / total_trades) * 100
                
                # 計算平均收益率
                avg_profit = sum(t['profit_rate'] for t in profit_trades) / len(profit_trades) if profit_trades else 0
                avg_loss = sum(t['profit_rate'] for t in [t for t in range_trades if t['profit_rate'] < 0]) / len([t for t in range_trades if t['profit_rate'] < 0]) if [t for t in range_trades if t['profit_rate'] < 0] else 0
                
                # 評分計算
                # 1. 勝率分數 (0-35分)
                win_rate_score = (win_rate / 100) * win_rate_weight * 100
                
                # 2. 收益率分數 (0-35分)
                profit_ratio = abs(avg_profit / avg_loss) if avg_loss != 0 else avg_profit
                profit_ratio_score = min(profit_ratio, 3) / 3 * profit_ratio_weight * 100
                
                # 3. 交易數量分數 (0-30分)
                trade_count_score = min(total_trades / 20, 1) * trade_count_weight * 100
                
                # 總分
                total_score = win_rate_score + profit_ratio_score + trade_count_score
                
                # 更新最佳區間
                if total_score > best_score:
                    best_score = total_score
                    best_range = (lower_cap, upper_cap)
                    best_metrics = {
                        '交易次數': total_trades,
                        '勝率': win_rate,
                        '平均獲利': avg_profit * 100,
                        '平均虧損': avg_loss * 100,
                        '盈虧比': abs(avg_profit / avg_loss) if avg_loss != 0 else float('inf'),
                        '綜合得分': total_score,
                        '市值下限': lower_cap,
                        '市值上限': upper_cap,
                        '區間內代幣數': len(set(t['market_cap'] for t in range_trades))
                    }
        
        if best_range is None:
            return {
                'success': False,
                'message': '未找到符合條件的市值區間'
            }
            
        return {
            'success': True,
            'best_range': best_range,
            'metrics': best_metrics,
            'all_market_caps': sorted(set(t['market_cap'] for t in token_trades)),
            'market_cap_stats': {
                'min': min_market_cap,
                'max': max_market_cap,
                'total_unique_caps': len(set(t['market_cap'] for t in token_trades))
            }
        }

    #====== 進階分析方法 ======
    def advanced_analysis(self, wallet_address, transactions, token_analysis):
        """執行進階分析，包括固定市值區間和動態最佳區間分析"""
        # 準備數據：收集所有代幣交易，並確保不重複
        token_trades = {}  # 使用字典來避免重複
        for token_address, token_data in {**token_analysis['profit_tokens'], **token_analysis['loss_tokens']}.items():
            market_cap = token_data.get('market_cap')
            if market_cap:  # 只收集有市值數據的交易
                for tx in transactions:
                    if ('token' in tx and 'address' in tx['token'] and 
                        tx['token']['address'] == token_address):
                        # 使用時間戳和代幣地址作為唯一鍵來避免重複
                        trade_key = (tx['時間戳'], token_address)
                        token_trades[trade_key] = {
                            'market_cap': market_cap,
                            'profit_rate': tx['收益率']
                        }
        
        # 將字典轉換為列表
        token_trades = list(token_trades.values())

        if not token_trades:
            return {
                'fixed_ranges': {},
                'dynamic_range': {'range': None, 'metrics': None},
                'market_stats': None
            }

        # 固定市值區間定義
        market_cap_ranges = [
            (0, 8888),          # $0K-$8.888K
            (0, 10000),         # $0K-$10K
            (5000, 10000),      # $5K-$10K
            (5000, 15000),      # $5K-$15K
            (15000, 30000),     # $15K-$30K
            (30000, 50000),     # $30K-$50K
            (50000, 100000),    # $50K-$100K
            (100000, float('inf'))  # >$100K
        ]

        # 評分參數設置
        MIN_TRADES = 5  # 最小交易數量要求
        win_rate_weight = 0.35      # 勝率權重
        profit_ratio_weight = 0.35  # 收益率權重
        trade_count_weight = 0.30   # 交易數量權重

        # 分析固定區間
        fixed_ranges = {}
        for low, high in market_cap_ranges:
            # 設置區間名稱
            if high == 8888:
                range_name = f"${low/1000:.0f}K-$8.888K"
            elif high == float('inf'):
                range_name = f">${low/1000:.0f}K"
            else:
                range_name = f"${low/1000:.0f}K-${high/1000:.0f}K"

            # 篩選區間內的交易
            range_trades = [t for t in token_trades if low <= t['market_cap'] < high]
            
            if len(range_trades) >= MIN_TRADES:
                profit_trades = [t for t in range_trades if t['profit_rate'] > 0]
                loss_trades = [t for t in range_trades if t['profit_rate'] < 0]
                
                total_trades = len(range_trades)
                win_rate = (len(profit_trades) / total_trades) * 100
                
                # 計算該區間的5倍爆擊率
                five_x_trades = [t for t in range_trades if t['profit_rate'] >= 5.0]  # 收益率400%即為5倍
                five_x_rate = (len(five_x_trades) / total_trades) * 100 if total_trades > 0 else 0
                
                # 計算平均收益率
                avg_profit = (sum(t['profit_rate'] for t in profit_trades) / len(profit_trades) * 100) if profit_trades else 0
                avg_loss = (sum(t['profit_rate'] for t in loss_trades) / len(loss_trades) * 100) if loss_trades else 0
                
                # 評分計算
                win_rate_score = (win_rate / 100) * win_rate_weight * 100
                profit_ratio = abs(avg_profit / avg_loss) if avg_loss != 0 else avg_profit
                profit_ratio_score = min(profit_ratio, 3) / 3 * profit_ratio_weight * 100
                trade_count_score = min(total_trades / 20, 1) * trade_count_weight * 100
                total_score = win_rate_score + profit_ratio_score + trade_count_score
                
                fixed_ranges[range_name] = {
                    '交易次數': total_trades,
                    '勝率': win_rate,
                    '平均獲利': avg_profit,
                    '平均虧損': avg_loss,
                    '5倍爆擊率': five_x_rate,  # 確保這個字段存在
                    '綜合得分': total_score
                }

        # 動態最佳區間分析
        best_score = 0
        best_range = None
        best_metrics = None
        
        # 獲取所有唯一的市值並排序
        all_market_caps = sorted(set(t['market_cap'] for t in token_trades))
        
        # 使用滑動窗口尋找最佳區間
        for i in range(len(all_market_caps)):
            for j in range(i + MIN_TRADES - 1, len(all_market_caps)):
                lower_cap = all_market_caps[i]
                upper_cap = all_market_caps[j]
                
                # 獲取區間內的所有交易
                range_trades = [t for t in token_trades 
                            if lower_cap <= t['market_cap'] <= upper_cap]
                
                if len(range_trades) < MIN_TRADES:
                    continue
                    
                # 計算區間指標
                total_trades = len(range_trades)
                profit_trades = [t for t in range_trades if t['profit_rate'] > 0]
                loss_trades = [t for t in range_trades if t['profit_rate'] < 0]
                
                if not profit_trades:  # 如果沒有獲利交易，跳過此區間
                    continue
                
                win_rate = (len(profit_trades) / total_trades) * 100
                
                # 計算平均收益率
                avg_profit = (sum(t['profit_rate'] for t in profit_trades) / len(profit_trades) * 100)
                avg_loss = (sum(t['profit_rate'] for t in loss_trades) / len(loss_trades) * 100) if loss_trades else 0
                
                # 計算盈虧比
                profit_ratio = abs(avg_profit / avg_loss) if avg_loss != 0 else float('inf')
                
                # 評分計算
                win_rate_score = (win_rate / 100) * win_rate_weight * 100
                profit_ratio_score = min(profit_ratio, 3) / 3 * profit_ratio_weight * 100
                trade_count_score = min(total_trades / 20, 1) * trade_count_weight * 100
                
                total_score = win_rate_score + profit_ratio_score + trade_count_score
                
                if total_score > best_score:
                    best_score = total_score
                    best_range = (lower_cap, upper_cap)
                    best_metrics = {
                        '交易次數': total_trades,
                        '勝率': win_rate,
                        '平均獲利': avg_profit,
                        '平均虧損': avg_loss,
                        '盈虧比': profit_ratio,
                        '區間內代幣數': len(set(t['market_cap'] for t in range_trades)),
                        '綜合得分': total_score
                    }

        # 統計資料
        market_stats = {
            'min': min(t['market_cap'] for t in token_trades),
            'max': max(t['market_cap'] for t in token_trades),
            'total_unique_caps': len(all_market_caps)
        }

        # 返回完整分析結果
        return {
            'fixed_ranges': fixed_ranges,
            'dynamic_range': {
                'range': best_range,
                'metrics': best_metrics
            },
            'market_stats': market_stats
        }

"""  
#====== 主程序 ====== 
def main():
   analyzer = WalletAnalyzer()
   
   while True:
       print("\n請輸入要分析的錢包地址（每行一個，輸入空行結束）：")
       wallet_addresses = []
       while True:
           address = input().strip()
           if not address:
               break
           wallet_addresses.append(address)
           
       if not wallet_addresses:
           print("未輸入錢包地址，程序結束")
           break
           
       for wallet_address in wallet_addresses:
           try:
               transactions = analyzer.fetch_transactions(wallet_address)
               if transactions:
                   analysis = analyzer.analyze_transactions(transactions)
                   token_analysis = analyzer.analyze_tokens_by_profit(wallet_address, transactions)
                   advanced_results = analyzer.advanced_analysis(wallet_address, transactions, token_analysis)
                   
                   print("\n正在生成Excel報告...")
                   analyzer.save_to_excel(wallet_address, transactions, analysis, token_analysis, advanced_results)
                   
                   # 基本分析摘要
                   print("\n基本分析結果摘要：")
                   print("=" * 50)
                   for key, value in analysis.items():
                       if key in ['獲利次數', '虧損次數', '總交易次數']:
                           print(f"{key}: {value}")
                       elif key == '盈虧比':
                           print(f"{key}: {value:.2f}")
                       elif '率' in key or '利' in key or '損' in key:
                           print(f"{key}: {value:.2f}%")
                   print(f"Bot機率: {token_analysis['quick_trade_ratio']:.2f}%")
                   print("=" * 50)

                   # 動態最佳區間分析結果
                   if advanced_results['dynamic_range']['range']:
                       best_range = advanced_results['dynamic_range']['range']
                       metrics = advanced_results['dynamic_range']['metrics']
                       print("\n動態最佳區間分析結果：")
                       print("=" * 50)
                       print(f"最佳交易市值區間: ${best_range[0]:,.0f} - ${best_range[1]:,.0f}")
                       print(f"區間內交易次數: {metrics['交易次數']}")
                       print(f"勝率: {metrics['勝率']:.2f}%")
                       print(f"平均獲利: {metrics['平均獲利']:.2f}%")
                       print(f"平均虧損: {metrics['平均虧損']:.2f}%")
                       
                       # 修正盈虧比的顯示邏輯
                       if metrics['盈虧比'] == float('inf'):
                           print("盈虧比: inf")
                       else:
                           print(f"盈虧比: {metrics['盈虧比']:.2f}")
                       
                       print(f"區間內代幣數: {metrics['區間內代幣數']}")
                       print(f"綜合得分: {metrics['綜合得分']:.2f}")
                       print("=" * 50)
               else:
                   print(f"無法獲取錢包 {wallet_address} 的交易記錄")
           
           except Exception as e:
               print(f"分析錢包 {wallet_address} 時發生錯誤: {str(e)}")
               import traceback
               traceback.print_exc()
               print("等待60秒後繼續下一個錢包...")
               time.sleep(60)

if __name__ == "__main__":
   main()
"""   
